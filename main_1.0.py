import os
import openpyxl
from tkinter import *
from tkinter import filedialog as fd
from tkinter.scrolledtext import ScrolledText

window = Tk()
window.geometry("640x450+650+150")
window.title("Excel Data Extraction and Structuring Script")
window.resizable(False, False)

Label(text='(c) S. Nikulin').grid(row=3, padx=0, pady=0, sticky='w' )

bg1='gainsboro' 
bg2='rosybrown'
bg3='darkgray' 

up_frame  =  Frame(window,  width=300,  height=  400,  bg=bg1)
up_frame.grid(row=0,  column=0,  padx=5,  pady=5, sticky='w'+'e'+'n'+'s')
down_frame  =  Frame(window,  width=300,  height= 420,  bg=bg2)
down_frame.grid(row=2,  column=0,  padx=5,  pady=5, sticky='w'+'e'+'n'+'s')

directory = ""

def open_directory():
    global directory 
    directory = fd.askdirectory()
    report_list = os.listdir(directory)
    st.insert(END, "Processing files in directory:" + str(directory) + '\n')
    st.insert(END, "Number of files:" + str(len(report_list)) + '\n')


column_1 = Entry(up_frame, width=5)
column_1.grid(row=3, column=2, padx=5, pady=5)
row_1 = Entry(up_frame, width=5)
row_1.grid(row=3, column=3, padx=5, pady=5)

column_2 = Entry(up_frame, width=5)
column_2.grid(row=4, column=2, padx=5, pady=5)
row_2 = Entry(up_frame, width=5)
row_2.grid(row=4, column=3, padx=5, pady=5)

column_3 = Entry(up_frame, width=5)
column_3.grid(row=5, column=2, padx=5, pady=5)
row_3 = Entry(up_frame, width=5)
row_3.grid(row=5, column=3, padx=5, pady=5)

column_4 = Entry(up_frame, width=5)
column_4.grid(row=6, column=2, padx=5, pady=5)
row_4 = Entry(up_frame, width=5)
row_4.grid(row=6, column=3, padx=5, pady=5)

column_5 = Entry(up_frame, width=5)
column_5.grid(row=7, column=2, padx=5, pady=5)
row_5 = Entry(up_frame, width=5)
row_5.grid(row=7, column=3, padx=5, pady=5)

def get_value():
    position_col_1 = column_1.get()
    position_row_1 = row_1.get()

    position_col_2 = column_2.get()
    position_row_2 = row_2.get()

    position_col_3 = column_3.get()
    position_row_3 = row_3.get()

    position_col_4 = column_4.get()
    position_row_4 = row_4.get()

    position_col_5 = column_5.get()
    position_row_5 = row_5.get()

    global directory
    report_list = os.listdir(directory)
    cell_values = []
    for report_file in report_list:
        if report_file.endswith((".xlsx", ".xls", ".xlsm", ".xlsb", ".xltx", ".xltm")):
            file_path = os.path.join(directory, report_file)
            st.insert(END, "Processing file:" + str(file_path) + '\n')
            print("Processing file:", file_path)
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            cell_value_1 = sheet[position_col_1 + position_row_1].value
            cell_value_2 = sheet[position_col_2 + position_row_2].value
            cell_value_3 = sheet[position_col_3 + position_row_3].value
            cell_value_4 = sheet[position_col_4 + position_row_4].value
            cell_value_5 = sheet[position_col_5 + position_row_5].value

            print("Cell value 1:", cell_value_1)
            st.insert(END, "Cell value 1:" + str(cell_value_1) + '\n')
            print("Cell value 2:", cell_value_2)
            st.insert(END, "Cell value 2:" + str(cell_value_2) + '\n')
            print("Cell value 3:", cell_value_3)
            st.insert(END, "Cell value 3:" + str(cell_value_3) + '\n')
            print("Cell value 4:", cell_value_4)
            st.insert(END, "Cell value 4:" + str(cell_value_4) + '\n')
            print("Cell value 5:", cell_value_5)
            st.insert(END, "Cell value 5:" + str(cell_value_5) + '\n')
            cell_values.append((cell_value_1, cell_value_2, cell_value_3, cell_value_4, cell_value_5))

    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active
    for row, values in enumerate(cell_values, start=1):
        for col, value in enumerate(values, start=1):
            output_sheet.cell(row=row, column=col, value=value)
    output_wb.save("output.xlsx")
    st.insert(END, "Cell values written to output.xlsx" + '\n')
    print("Cell values written to output.xlsx")

language = "russian"

def switch_language():
    global language 
    if language == "russian":
        language = "english"
    else:
        language = "russian"
    update_labels()

def update_labels():
    global language  # Use the global 'language' variable
    if language == "russian":
        Label(up_frame, text='Укажите путь к папке', bg=bg1, width=25).grid(row=1, column=1)
        Label(up_frame, text='Позиция по вертикали', bg=bg1, width=25).grid(row=2, column=2)
        Label(up_frame, text='Позиция по горизонтали', bg=bg1, width=25).grid(row=2, column=3)
        Label(up_frame, text='Введите столбец и строку ячеек:', bg=bg1, width=25).grid(row=2, column=1)
        Button(up_frame, text="Путь", command=open_directory, width=15).grid(row=1, column=2)
        Button(down_frame, text='Сформировать таблицу данных', command=get_value, width=25).grid(row=1, column=1, sticky='w', padx=5, pady=5)
        Button(up_frame, text='En', command=switch_language, bg=bg3, width=5). grid(row=1, column=4, sticky='e') 
    else:
        Label(up_frame, text='Select folder path', bg=bg1, width=25).grid(row=1, column=1)
        Label(up_frame, text='Vertical Position', bg=bg1, width=25).grid(row=2, column=2)
        Label(up_frame, text='Horizontal Position', bg=bg1, width=25).grid(row=2, column=3)
        Label(up_frame, text='Enter column and row for cells:', bg=bg1, width=25).grid(row=2, column=1)
        Button(up_frame, text="Path", command=open_directory, width=15).grid(row=1, column=2)
        Button(down_frame, text='Generate data table', command=get_value, width=25).grid(row=1, column=1, sticky='w', padx=5, pady=5)
        Button(up_frame, text='Ru', command=switch_language, bg=bg3, width=5). grid(row=1, column=4, sticky='e') 

Label(up_frame, text='Укажите путь к папке', bg=bg1, width=25).grid(row=1, column=1)
Label(up_frame, text='Позиция по вертикали', bg=bg1, width=25).grid(row=2, column=2)
Label(up_frame, text='Позиция по горизонтали', bg=bg1, width=25).grid(row=2, column=3)
Label(up_frame, text='Введите столбец и строку ячеек:', bg=bg1, width=25).grid(row=2, column=1)
Button(up_frame, text="Путь", command=open_directory, width=15).grid(row=1, column=2)
Button(down_frame, text='Сформировать таблицу данных', command=get_value, width=25).grid(row=1, column=1, sticky='w', padx=5, pady=5)
Button(up_frame, text='En', command=switch_language, bg=bg3, width=5). grid(row=1, column=4, sticky='e') 

#описание процесса
st = ScrolledText(down_frame, width=85,  height=10, bd=1.5, font = 'Arial 10')
st.grid(row=2, column=1, padx=5, pady=5, sticky='w'+'e'+'n'+'s')

window.mainloop()
